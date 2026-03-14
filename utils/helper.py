"""
Helper utilities
"""
import re
import discord
from utils.logger import get_logger
from typing import Optional, Dict, Any, List, Tuple
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import aiohttp


logger = get_logger()


def sanitize_code_content(content: str) -> str:
    """
    Sanitize code content for analysis.
    
    Args:
        content: Raw file content
        
    Returns:
        Sanitized content
    """
    if not isinstance(content, str):
        return ""
        
    # Truncate large files
    content = content[:5000] #todo add from config
    
    # Remove Unicode control characters
    content = re.sub(r'[\u0000-\u0008\u000B-\u000C\u000E-\u001F\uD800-\uDFFF]', '', content)
    
    # Normalize newlines
    content = content.replace('\r\n', '\n').replace('\r', '\n')
    
    return content

def get_file_extension(filename: str) -> str:
    """
    Get language identifier from file extension.
    
    Args:
        filename: Filename including extension
        
    Returns:
        Language name for syntax highlighting
    """
    # Fast extension extraction
    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    
    # Language mapping
    language_map = {
        'js': 'javascript',
        'ts': 'typescript',
        'py': 'python',
        'java': 'java',
        'go': 'go',
        'rs': 'rust',
        'cpp': 'cpp',
        'c': 'c',
        'jsx': 'javascript',
        'tsx': 'typescript',
        'vue': 'vue',
        'php': 'php',
        'rb': 'ruby',
        'sol': 'solidity',
        'cs': 'csharp',
        'html': 'html',
        'css': 'css',
        'scss': 'scss',
        'md': 'markdown',
        'json': 'json',
        'yml': 'yaml',
        'yaml': 'yaml'
    }
    
    return language_map.get(ext, ext)

def safe_add_field(content):
    """Truncate content to fit Discord's 1024 character limit"""
    if len(content) > 1024:
        return content[:1020] + "..."
    return content

async def get_token_metadata_and_links(uri, dev_address=None):
    """
    Reusable method to fetch token metadata from URI and extract links
    Returns: (image_url, links_list)
    """
    try:
        if not uri:
            return None, []
        
        # Fetch metadata from URI
        metadata = await fetch_token_metadata_from_uri(uri)
        if not metadata:
            return None, []
        
        # Extract image
        image_url = metadata.get('image')
        
        # Build links
        links = []
        
        # Add dev link if provided
        if dev_address:
            links.append({"label": "Dev", "url": f"https://solscan.io/account/{dev_address}"})
        
        # Social media mapping
        social_fields = {
            'twitter': 'Twitter',
            'telegram': 'Telegram', 
            'discord': 'Discord',
            'youtube': 'YouTube',
            'reddit': 'Reddit',
            'github': 'GitHub',
            'medium': 'Medium',
            'instagram': 'Instagram'
        }
        
        # Add direct social fields
        for field, label in social_fields.items():
            if field in metadata and metadata[field]:
                links.append({"label": label, "url": metadata[field]})
        
        # Check website field for socials or general website
        website = metadata.get('website') or metadata.get('external_url')
        if website:
            social_patterns = {
                ('twitter.com', 'x.com'): 'Twitter',
                ('t.me',): 'Telegram',
                ('discord.gg', 'discord.com', 'discord.app'): 'Discord',
                ('youtube.com', 'youtu.be'): 'YouTube',
                ('instagram.com',): 'Instagram',
                ('tiktok.com',): 'TikTok',
                ('reddit.com',): 'Reddit',
                ('github.com',): 'GitHub',
                ('medium.com',): 'Medium'
            }
            
            social_found = False
            for patterns, label in social_patterns.items():
                if any(pattern in website for pattern in patterns):
                    links.append({"label": label, "url": website})
                    social_found = True
                    break
            
            if not social_found:
                links.append({"label": "Web", "url": website})
        
        return image_url, links
        
    except Exception as e:
        logger.error(f"Error fetching metadata from URI {uri}: {e}")
        return None, []

async def fetch_token_metadata_from_uri(uri: str) -> Optional[dict]:
    """Fetch token metadata from URI and return full metadata"""
    if not uri:
        return None
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(uri, timeout=aiohttp.ClientTimeout(total=10)) as response:
                if response.status == 200:
                    metadata = await response.json()
                    return metadata
    except Exception as e:
        logger.debug(f"Error fetching token metadata from {uri}: {e}")
    
    return None

async def get_webhook_info(webhook_url: str, bot) -> Optional[Dict]:
    try:
        parts = webhook_url.rstrip("/").split("/")
        webhook_id = int(parts[-2])
        webhook = await bot.fetch_webhook(webhook_id)

        guild_id = webhook.guild_id
        channel_id = webhook.channel_id
        bot_in_guild = False
        channel_name = "Unknown (bot not in server)"

        if guild_id:
            for g in bot.guilds:
                if g.id == guild_id:
                    bot_in_guild = True
                    try:
                        channel = g.get_channel(channel_id) or await g.fetch_channel(channel_id)
                        channel_name = channel.name
                    except Exception:
                        channel_name = "Unknown"
                    break

        return {
            "guild_id": guild_id,
            "channel_id": channel_id,
            "channel_name": channel_name,
            "bot_in_guild": bot_in_guild
        }
    except Exception as e:
        logger.error(f"Failed to fetch webhook info: {e}", exc_info=True)
        return None

async def fetch_channel_global(bot, channel_id: int) -> Optional[discord.abc.GuildChannel]:
    for guild in bot.guilds:
        try:
            channel = guild.get_channel(channel_id) or await guild.fetch_channel(channel_id)
            if channel:
                return channel
        except Exception:
            continue
    return None

async def get_thread_name(thread_id: str, bot) -> Optional[str]:
    try:
        tid = int(thread_id)
    except ValueError:
        return None

    channel = await fetch_channel_global(bot, tid)
    if not channel:
        return None

    try:
        if isinstance(channel, discord.Thread):
            parent_name = channel.parent.name if channel.parent else "Unknown"
            return f"🧵 {channel.name} (in #{parent_name})"
        if isinstance(channel, discord.ForumChannel):
            return f"📋 {channel.name}"
        return f"#{channel.name}"
    except Exception as e:
        logger.error(f"Failed to format thread/channel name: {e}", exc_info=True)
        return None

def parse_market_cap_value(value: str) -> Optional[float]:
    """
    Parse market cap value from string format
    
    Args:
        value: Market cap string (e.g., "50k", "2.5m", "1b", "250000")
    
    Returns:
        float: Parsed market cap value
        
    Raises:
        ValueError: If value format is invalid
    """
    if not value:
        return None
    
    # Remove spaces and convert to lowercase
    value = value.strip().lower()
    
    # Handle direct numeric values
    if value.replace('.', '').replace(',', '').isdigit():
        return float(value.replace(',', ''))
    
    # Regular expression to match number + suffix
    pattern = r'^(\d+(?:\.\d+)?)\s*([kmb]?)$'
    match = re.match(pattern, value)
    
    if not match:
        raise ValueError(f"Invalid market cap format: {value}")
    
    number = float(match.group(1))
    suffix = match.group(2)
    
    # Apply multipliers
    multipliers = {
        'k': 1_000,
        'm': 1_000_000,
        'b': 1_000_000_000
    }
    
    multiplier = multipliers.get(suffix, 1)
    return number * multiplier

def calculate_mc_range(target_mc: float, cutoff: Optional[float]) -> tuple[float, float]:
    """
    Calculate market cap range based on target and cutoff
    
    Args:
        target_mc: Target market cap value
        cutoff: Optional cutoff value to create range
        
    Returns:
        tuple: (min_mc, max_mc)
    """
    if cutoff is None:
        # Return exact value for precise matching
        return (target_mc, target_mc)
    
    # Create range around target with cutoff
    min_mc = max(0, target_mc - cutoff)
    max_mc = target_mc + cutoff
    
    return (min_mc, max_mc)

def get_auto_tolerance(target_mc: float) -> float:
    """
    Calculate automatic tolerance based on target MC value
    
    Args:
        target_mc: Target market cap value
        
    Returns:
        float: Tolerance value
    """
    if target_mc >= 1_000_000_000:  # 1B+
        return target_mc * 0.001  # 0.1%
    elif target_mc >= 100_000_000:  # 100M+
        return target_mc * 0.002  # 0.2%
    elif target_mc >= 10_000_000:  # 10M+
        return target_mc * 0.005  # 0.5%
    elif target_mc >= 1_000_000:  # 1M+
        return target_mc * 0.01  # 1%
    elif target_mc >= 100_000:  # 100K+
        return target_mc * 0.02  # 2%
    else:  # Under 100K
        return target_mc * 0.05  # 5%

def parse_buy_amount_value(value: str) -> Optional[Dict[str, Any]]:
    """
    Parse buy amount value from string format
    
    Args:
        value: Buy amount string (e.g., "2 sol", "$150", "1.5 solana", "300 usd")
        
    Returns:
        Dict with 'amount', 'currency', 'original' or None
        
    Raises:
        ValueError: If value format is invalid
    """
    if not value:
        return None
    
    # Remove extra spaces and convert to lowercase
    value = value.strip().lower()
    original_value = value
    
    # Check for SOL variants
    sol_patterns = [
        r'^(\d+(?:\.\d+)?)\s*sol(?:ana)?$',  # "2 sol", "2.5 solana"
        r'^sol(?:ana)?\s*(\d+(?:\.\d+)?)$',  # "sol 2", "solana 2.5"
    ]
    
    for pattern in sol_patterns:
        match = re.match(pattern, value)
        if match:
            amount = float(match.group(1))
            return {
                'amount': amount,
                'currency': 'SOL',
                'original': original_value
            }
    
    # Check for USD variants
    usd_patterns = [
        r'^\$(\d+(?:\.\d+)?)$',  # "$150"
        r'^(\d+(?:\.\d+)?)\$$',  # "150$"
        r'^(\d+(?:\.\d+)?)\s*usd$',  # "150 usd"
        r'^(\d+(?:\.\d+)?)\s*dollars?$',  # "150 dollar", "150 dollars"
    ]
    
    for pattern in usd_patterns:
        match = re.match(pattern, value)
        if match:
            amount = float(match.group(1))
            return {
                'amount': amount,
                'currency': 'USD',
                'original': original_value
            }
    
    # Check for plain numbers
    number_match = re.match(r'^(\d+(?:\.\d+)?)$', value)
    if number_match:
        amount = float(number_match.group(1))
        # If less than 10, assume SOL, otherwise USD
        currency = 'SOL' if amount < 10 else 'USD'
        return {
            'amount': amount,
            'currency': currency,
            'original': original_value
        }
    
    raise ValueError(f"Invalid buy amount format: {value}")

def get_buy_amount_tolerance(amount: float, currency: str) -> float:
    """
    Calculate tolerance for buy amount filtering
    
    Args:
        amount: Buy amount value
        currency: Currency type ('SOL' or 'USD')
        
    Returns:
        float: Tolerance value
    """
    if currency == 'SOL':
        if amount >= 10:
            return amount * 0.15  # 15% tolerance for large SOL amounts
        elif amount >= 1:
            return amount * 0.25  # 25% tolerance for medium SOL amounts
        else:
            return amount * 0.4   # 40% tolerance for small SOL amounts
    else:  # USD
        if amount >= 10000:
            return amount * 0.1   # 10% tolerance for large USD amounts
        elif amount >= 1000:
            return amount * 0.15  # 15% tolerance for medium USD amounts
        elif amount >= 100:
            return amount * 0.25  # 25% tolerance for smaller USD amounts
        else:
            return amount * 0.4   # 40% tolerance for very small USD amounts

def get_first_candle_info(candlesticks: List[Dict]) -> Tuple[Optional[int], Optional[int]]:
    """
    Get the first trade price and timestamp from candlesticks
    
    Args:
        candlesticks: List of candlestick data
        
    Returns:
        Tuple of (first_price in cents, first_timestamp) or (None, None)
    """
    if not candlesticks:
        return None, None
    
    for candle in candlesticks:
        # Find first candle with volume (actual trades)
        if candle.get("volume", 0) > 0 and candle.get("price", {}).get("close") is not None:
            price = candle["price"]["close"]
            timestamp = candle.get("end_period_ts")
            return price, timestamp
    
    return None, None

def generate_candlestick_excel(markets: List[Dict], series_ticker: str, 
                               get_candlesticks_func) -> Optional[bytes]:
    """
    Generate Excel file with detailed candlestick data for all markets
    
    Args:
        markets: List of market data
        series_ticker: Series ticker for API calls
        get_candlesticks_func: Async function to fetch candlesticks
        
    Returns:
        Excel file bytes or None
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Market Candlesticks"
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Write header
        headers = [
            "Market Name",
            "Market Ticker",
            "Timestamp",
            "Time (UTC)",
            "Open Price (¢)",
            "High Price (¢)",
            "Low Price (¢)",
            "Close Price (¢)",
            "Volume",
            "Open Interest",
            "Yes Bid (¢)",
            "Yes Ask (¢)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        # Set column widths
        column_widths = [20, 30, 15, 22, 15, 15, 15, 15, 12, 15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        current_row = 2
        current_ts = int(datetime.now().timestamp())
        
        # This needs to be called from an async context
        # We'll return a coroutine that needs to be awaited
        return wb, current_row, current_ts
        
    except Exception as e:
        logger.error(f"Error creating Excel workbook: {e}")
        return None

async def populate_excel_data(wb, start_row: int, markets: List[Dict], 
                              series_ticker: str, get_candlesticks_func) -> Optional[bytes]:
    """
    Populate Excel workbook with candlestick data
    
    Args:
        wb: Workbook object
        start_row: Starting row for data
        markets: List of market data
        series_ticker: Series ticker
        get_candlesticks_func: Async function to fetch candlesticks
        
    Returns:
        Excel file bytes or None
    """
    try:
        ws = wb.active
        current_row = start_row
        current_ts = int(datetime.now().timestamp())
        
        # Process each market
        for market in markets:
            market_name = market.get("yes_sub_title", "Unknown")
            ticker = market.get("ticker", "")
            
            if not ticker or not series_ticker:
                continue
            
            open_time = market.get("open_time")
            if not open_time:
                continue
            
            try:
                start_ts = int(datetime.fromisoformat(open_time.replace('Z', '+00:00')).timestamp())
                candlesticks = await get_candlesticks_func(series_ticker, ticker, start_ts, current_ts)
                
                if not candlesticks:
                    continue
                
                # Write each candlestick
                for candle in candlesticks:
                    timestamp = candle.get("end_period_ts")
                    if not timestamp:
                        continue
                    
                    time_utc = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
                    volume = candle.get("volume", 0)
                    open_interest = candle.get("open_interest", 0)
                    
                    price_data = candle.get("price", {})
                    open_price = price_data.get("open") if price_data.get("open") is not None else ""
                    high_price = price_data.get("high") if price_data.get("high") is not None else ""
                    low_price = price_data.get("low") if price_data.get("low") is not None else ""
                    close_price = price_data.get("close") if price_data.get("close") is not None else ""
                    
                    yes_bid_data = candle.get("yes_bid", {})
                    yes_ask_data = candle.get("yes_ask", {})
                    yes_bid = yes_bid_data.get("close") if yes_bid_data.get("close") is not None else ""
                    yes_ask = yes_ask_data.get("close") if yes_ask_data.get("close") is not None else ""
                    
                    row_data = [
                        market_name,
                        ticker,
                        timestamp,
                        time_utc,
                        open_price,
                        high_price,
                        low_price,
                        close_price,
                        volume,
                        open_interest,
                        yes_bid,
                        yes_ask
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=current_row, column=col, value=value)
                    
                    current_row += 1
                    
            except Exception as e:
                logger.warning(f"Could not process candlesticks for {ticker}: {e}")
                continue
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Error populating Excel data: {e}")
        return None
    
async def resolve_user(
    bot: discord.Client,
    user_id: int,
    guild: discord.Guild,
    scope: str
) -> Optional[tuple[str, discord.Member | discord.User]]:
    """
    Try to resolve a user_id to a live Discord user/member object.
    Returns (display_string, user_obj) or None if ghost/deleted/banned.
    """
    try:
        if scope == "server":
            user = guild.get_member(user_id)
        else:
            user = bot.get_user(user_id)
            if user is None:
                user = await bot.fetch_user(user_id)

        if user is None:
            return None

        display = f"{user.display_name} (@{user.name})"
        return display, user

    except discord.NotFound:
        return None
    except Exception as e:
        logger.warning(f"Could not resolve user {user_id}: {e}")
        return None